from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session
import traceback
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import date, timedelta
import datetime
from functools import wraps
import sqlite3
import pytz
import json
import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER

def adapt_datetime(dt):
    return dt.isoformat()

def convert_datetime(val):
    return datetime.datetime.fromisoformat(val.decode())

sqlite3.register_adapter(datetime.datetime, adapt_datetime)
sqlite3.register_converter("TIMESTAMP", convert_datetime)

app = Flask(__name__)

@app.template_filter('to_local')
def to_local(dt, tz_name='Asia/Kuala_Lumpur'):
    if not dt:
        return ''
    if isinstance(dt, str):
        # If your timestamp is a string (e.g., from SQLite)
        try:
            dt = datetime.datetime.fromisoformat(dt)
        except ValueError:
            return dt
    local_tz = pytz.timezone(tz_name)
    if dt.tzinfo is None:
        dt = pytz.utc.localize(dt)
    local_dt = dt.astimezone(local_tz)
    return local_dt.strftime('%d-%m-%Y %H:%M:%S')


# Session Configuration - NO Flask-Session needed
app.config['SECRET_KEY'] = 'your-secret-key-change-in-production'
app.config['SESSION_COOKIE_NAME'] = 'ssm_session'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = False  # Set True if using HTTPS
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=1)

app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg', 'pdf', 'doc', 'docx'}


# Ensure upload directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Flask-Login setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.session_protection = 'strong'  # Add this for better security

# Database setup
DATABASE = 'weighing_scales.db'

import sqlite3
import datetime

def get_db():
    # Register adapters (Python → SQLite)
    sqlite3.register_adapter(datetime.date, lambda d: d.isoformat())
    sqlite3.register_adapter(datetime.datetime, lambda dt: dt.isoformat())

    # Register converters (SQLite → Python)
    sqlite3.register_converter("DATE", lambda s: datetime.date.fromisoformat(s.decode()))
    sqlite3.register_converter("DATETIME", lambda s: datetime.datetime.fromisoformat(s.decode()))

    # Connect with type detection enabled
    conn = sqlite3.connect(
        DATABASE,
        detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES
    )

    conn.row_factory = sqlite3.Row
    return conn

    
def init_db():
    with app.app_context():
        db = get_db()
        db.executescript('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                full_name TEXT NOT NULL,
                email TEXT,
                role TEXT NOT NULL CHECK(role IN ('admin', 'technician', 'support')),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            
            CREATE TABLE IF NOT EXISTS logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                username TEXT NOT NULL,
                action TEXT NOT NULL,
                table_name TEXT NOT NULL,
                record_id INTEGER,
                old_data TEXT,
                new_data TEXT,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS supermarkets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                head_office_address TEXT,
                contact_person TEXT,
                phone TEXT,
                email TEXT,
                remarks TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS branches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                supermarket_id INTEGER NOT NULL,
                branch_name TEXT NOT NULL,
                branch_code TEXT,
                branch_region TEXT,
                state TEXT NOT NULL,
                app_version TEXT NOT NULL,
                address TEXT,
                contact_person TEXT,
                phone TEXT,
                branch_total TEXT,
                opening_date DATE,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (supermarket_id) REFERENCES supermarkets(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS weighing_scales (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                supermarket_id INTEGER NOT NULL,
                branch_id INTEGER NOT NULL,
                brand TEXT NOT NULL,
                model TEXT NOT NULL,
                serial_number TEXT UNIQUE NOT NULL,
                firmware_number TEXT NOT NULL,
                installation_date DATE,
                ip_address TEXT,
                mac_address TEXT,
                anydesk_id TEXT,
                anydesk_password TEXT,
                weight_license_number TEXT,
                license_expiry_date DATE,
                maintenance_status TEXT CHECK(maintenance_status IN ('active', 'expired', 'pending')),
                technician_in_charge TEXT,
                remarks TEXT,
                document_path TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (supermarket_id) REFERENCES supermarkets(id) ON DELETE CASCADE,
                FOREIGN KEY (branch_id) REFERENCES branches(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS maintenance_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                scale_id INTEGER NOT NULL,
                technician_name TEXT NOT NULL,
                service_date DATE NOT NULL,
                issue_description TEXT,
                resolution TEXT,
                next_service_due DATE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (scale_id) REFERENCES weighing_scales(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS notifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                message TEXT NOT NULL,
                type TEXT CHECK(type IN ('license_expiry', 'maintenance_due', 'alert')),
                related_id INTEGER,
                is_read BOOLEAN DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        ''')
        
        # Check if logs table has user_id column, add it if it doesn't exist
        cursor = db.execute("PRAGMA table_info(logs)")
        columns = [column[1] for column in cursor.fetchall()]
        if 'user_id' not in columns:
            db.execute("ALTER TABLE logs ADD COLUMN user_id INTEGER NOT NULL DEFAULT 0")
            db.commit()
        
        # Check if branches table has branch_code column, add it if it doesn't exist
        cursor = db.execute("PRAGMA table_info(branches)")
        columns = [column[1] for column in cursor.fetchall()]
        if 'branch_code' not in columns:
            db.execute("ALTER TABLE branches ADD COLUMN branch_code TEXT")
            db.commit()
        
        # Create default admin user if not exists
        cursor = db.execute("SELECT * FROM users WHERE username = 'admin'")
        if not cursor.fetchone():
            hashed_password = generate_password_hash('admin123')
            db.execute(
                "INSERT INTO users (username, password, full_name, email, role) VALUES (?, ?, ?, ?, ?)",
                ('admin', hashed_password, 'System Administrator', 'admin@example.com', 'admin')
            )
        
        db.commit()
        db.close()

# User class for Flask-Login
class User(UserMixin):
    def __init__(self, id, username, full_name, email, role):
        self.id = id
        self.username = username
        self.full_name = full_name
        self.email = email
        self.role = role

@login_manager.user_loader
def load_user(user_id):
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    db.close()
    if user:
        return User(user['id'], user['username'], user['full_name'], user['email'], user['role'])
    return None

def get_changes(old_data, new_data):
    """
    Compare old_data and new_data dictionaries,
    return only fields that were actually updated (non-empty changes).
    """
    old_change = {}
    new_change = {}

    for key, new_value in new_data.items():
        old_value = old_data.get(key) if old_data else None

        # Convert None to empty string for comparison
        old_str = '' if old_value is None else str(old_value)
        new_str = '' if new_value is None else str(new_value)

        # Only log if value actually changed
        if old_str != new_str and new_str.strip() != '':
            old_change[key] = old_str
            new_change[key] = new_str

    if not old_change:
        return None, None
    return old_change, new_change


    
def log_action(action, table_name, record_id=None, old_data=None, new_data=None):
    if not current_user.is_authenticated:
        return

    # If old_data/new_data are dicts, convert to string of values
    if isinstance(old_data, dict):
        old_val = ', '.join(str(v) for v in old_data.values())
    else:
        old_val = str(old_data) if old_data else None

    if isinstance(new_data, dict):
        new_val = ', '.join(str(v) for v in new_data.values())
    else:
        new_val = str(new_data) if new_data else None

    db = get_db()
    db.execute(
        '''INSERT INTO logs (user_id, username, action, table_name, record_id, old_data, new_data)
           VALUES (?, ?, ?, ?, ?, ?, ?)''',
        (
            current_user.id,
            current_user.username,
            action,
            table_name,
            record_id,
            old_val,
            new_val
        )
    )
    db.commit()
    db.close()
    
# JSON file path
TICKETS_FILE = 'tickets.json'

@app.template_filter('format_date')
def format_date(date_str, format='%d.%m.%Y'):
    """Convert date string to formatted date"""
    try:
        if ' ' in date_str:
            # If it has time, split it
            date_str = date_str.split()[0]
        # Parse YYYY-MM-DD and format to DD-MM-YYYY
        date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d')
        return date_obj.strftime(format)
    except:
        return date_str

def load_tickets():
    """Load tickets from JSON file"""
    if os.path.exists(TICKETS_FILE):
        with open(TICKETS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def save_tickets(tickets):
    """Save tickets to JSON file"""
    with open(TICKETS_FILE, 'w', encoding='utf-8') as f:
        json.dump(tickets, f, indent=2, ensure_ascii=False)

def generate_ticket_number():
    """Generate unique ticket number"""
    tickets = load_tickets()
    if not tickets:
        return "TKT-2025-0001"
    
    # Get the last ticket number and increment
    last_ticket = tickets[-1]['ticket_no']
    year = datetime.datetime.now().year
    number = int(last_ticket.split('-')[-1]) + 1
    return f"TKT-{year}-{number:04d}"
    

# Role-based access decorator
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('You need administrator privileges to access this page.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# Utility functions
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def check_license_expiry():
    """Check for licenses expiring within 30 days and create notifications"""
    db = get_db()
    today = datetime.datetime.now().date()
    expiry_threshold = today + timedelta(days=30)
    
    scales = db.execute(
        """SELECT ws.id, ws.serial_number, ws.license_expiry_date, s.name as supermarket, b.branch_name
           FROM weighing_scales ws
           JOIN supermarkets s ON ws.supermarket_id = s.id
           JOIN branches b ON ws.branch_id = b.id
           WHERE ws.license_expiry_date <= ? AND ws.license_expiry_date >= ?""",
        (expiry_threshold, today)
    ).fetchall()
    
    for scale in scales:
        # Check if notification already exists
        existing = db.execute(
            "SELECT id FROM notifications WHERE related_id = ? AND type = 'license_expiry' AND is_read = 0",
            (scale['id'],)
        ).fetchone()
        
        if not existing:
            # scale['license_expiry_date'] is now already a datetime.date object!
            days_left = (scale['license_expiry_date'] - today).days     
            message = f"License for scale {scale['serial_number']} at {scale['supermarket']} - {scale['branch_name']} expires in {days_left} days"
            db.execute(
                "INSERT INTO notifications (title, message, type, related_id) VALUES (?, ?, ?, ?)",
                ('License Expiring Soon', message, 'license_expiry', scale['id'])
            )
    
    db.commit()
    db.close()

# Routes
@app.route('/')
@login_required
def index():
    return redirect(url_for('dashboard'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    # Clear any leftover flash messages before showing login page
    session.pop('_flashes', None)
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        db = get_db()
        user = db.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
        db.close()
        
        if user and check_password_hash(user['password'], password):
            # Check if user account is active
            try:
                is_active = user['is_active']
            except (KeyError, IndexError):
                is_active = 1  # Default to 1 (active) if column doesn't exist yet
            
            if is_active == 0:  # 0 = disabled, 1 = active
                flash('Your account has been disabled. Please contact the administrator.', 'danger')
                return redirect(url_for('login'))
            
            user_obj = User(user['id'], user['username'], user['full_name'], user['email'], user['role'])
            
            # CRITICAL: Make session non-permanent BEFORE login
            session.permanent = False
            login_user(user_obj, remember=False, fresh=True)
            
            # Log login action
            log_action('login', 'users', user['id'], None, user['username'])
            
            flash(f'Welcome back, {user["full_name"]}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    #log_action('logout', 'users', current_user.id, {'username': current_user.username}, None)
    
    # Log logout action
    #log_action('logout', 'users', current_user.id, None, current_user.username)
    
    logout_user()
    session.clear()
    session.pop('_flashes', None)
    flash('You have been logged out successfully', 'info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    check_license_expiry()  # Check for expiring licenses
    
    db = get_db()
    
    # Get statistics
    stats = {
        'total_supermarkets': db.execute("SELECT COUNT(*) as count FROM supermarkets").fetchone()['count'],
        'total_branches': db.execute("SELECT COUNT(*) as count FROM branches").fetchone()['count'],
        'total_scales': db.execute("SELECT COUNT(*) as count FROM weighing_scales").fetchone()['count'],
        'total_supermarkets': db.execute("SELECT COUNT(*) as count FROM supermarkets").fetchone()['count'],
        'total_branches': db.execute("SELECT COUNT(*) as count FROM branches").fetchone()['count'],
        'active_contracts': db.execute("SELECT COUNT(*) as count FROM weighing_scales WHERE maintenance_status = 'active'").fetchone()['count']
    }
    
    # Get all supermarkets for the dropdown
    supermarkets = db.execute("SELECT id, name FROM supermarkets ORDER BY name").fetchall()
    supermarkets = [dict(row) for row in supermarkets]
    
    # Get all branches for the dropdown
    branches = db.execute("SELECT id, branch_name FROM branches ORDER BY branch_name").fetchall()
    branches = [dict(row) for row in branches]
    
    # Get expiring licenses (within 30 days)
    today = datetime.datetime.now().date()
    expiry_threshold = today + timedelta(days=30)
    expiring_licenses = db.execute(
        """SELECT ws.*, s.name as supermarket, b.branch_name, b.state
           FROM weighing_scales ws
           JOIN supermarkets s ON ws.supermarket_id = s.id
           JOIN branches b ON ws.branch_id = b.id
           WHERE ws.license_expiry_date <= ? AND ws.license_expiry_date >= ?
           ORDER BY ws.license_expiry_date ASC
           LIMIT 10""",
        (expiry_threshold.isoformat(), today.isoformat())
    ).fetchall()
    
    expiring_licenses = [dict(row) for row in expiring_licenses]
    stats['expiring_licenses'] = len(expiring_licenses)
    
    # 👉 ADD THIS (CORRECT DAY CALCULATION)
    for scale in expiring_licenses:
        if scale['license_expiry_date']:
            expiry = scale['license_expiry_date']
            scale['days_left'] = (expiry - today).days
        else:
            scale['days_left'] = None
    
    # Installation per state
    state_data = db.execute(
        """SELECT b.state, COUNT(ws.id) as count
           FROM weighing_scales ws
           JOIN branches b ON ws.branch_id = b.id
           GROUP BY b.state
           ORDER BY count DESC"""
    ).fetchall()
    state_data = [dict(row) for row in state_data]
    
    # Maintenance status distribution
    maintenance_data = db.execute(
        """SELECT maintenance_status, COUNT(*) as count
           FROM weighing_scales
           GROUP BY maintenance_status"""
    ).fetchall()
    maintenance_data = [dict(row) for row in maintenance_data]
    
    # Latest maintenance records
    latest_maintenance = db.execute(
        """SELECT mr.*, ws.serial_number, s.name as supermarket, b.branch_name
           FROM maintenance_records mr
           JOIN weighing_scales ws ON mr.scale_id = ws.id
           JOIN supermarkets s ON ws.supermarket_id = s.id
           JOIN branches b ON ws.branch_id = b.id
           ORDER BY mr.service_date DESC
           LIMIT 5"""
    ).fetchall()
    latest_maintenance = [dict(row) for row in latest_maintenance]
    
    # Get unread notifications
    notifications = db.execute(
        "SELECT * FROM notifications WHERE is_read = 0 ORDER BY created_at DESC LIMIT 5"
    ).fetchall()
    notifications = [dict(row) for row in notifications]
    
    # Get recent logs (latest 20 entries)
    logs = db.execute(
        """SELECT id, user_id, username, action, table_name, record_id, old_data, new_data, timestamp
        FROM logs ORDER BY timestamp DESC LIMIT 5"""
    ).fetchall()

    logs = [dict(row) for row in logs]
    
    db.close()
    
    return render_template(
        'dashboard.html',
        stats=stats,
        supermarkets=supermarkets,
        branches=branches,
        expiring_licenses=expiring_licenses,
        state_data=state_data,
        maintenance_data=maintenance_data,
        latest_maintenance=latest_maintenance,
        notifications=notifications,
        logs=logs
    )

@app.route('/monitoring')
def monitoring():  # No login required!
    db = get_db()
    
    # Get search parameters
    search_supermarket = request.args.get('supermarket', '')
    search_branch = request.args.get('branch', '')
    search_serial = request.args.get('serial', '')
    search_app_version = request.args.get('app_version', '')
    search_state = request.args.get('state', '')
    
    # Build query with search filters
    query = """
        SELECT 
            ws.id,
            ws.serial_number,
            ws.firmware_number,
            ws.ip_address,
            ws.anydesk_id,
            b.branch_name,
            b.branch_code,
            b.branch_region,
            b.state,
            b.app_version,
            s.name as supermarket_name
        FROM weighing_scales ws
        INNER JOIN branches b ON ws.branch_id = b.id
        INNER JOIN supermarkets s ON ws.supermarket_id = s.id
        WHERE 1=1
    """
    
    params = []
    
    if search_supermarket:
        query += " AND s.name LIKE ?"
        params.append(f'%{search_supermarket}%')
    
    if  search_branch:
        query += """
            AND (
                LOWER(b.branch_name) LIKE LOWER(?)
                OR LOWER(b.branch_code) LIKE LOWER(?)
                OR LOWER(b.branch_region) LIKE LOWER(?)
            )
        """
        params.extend([f'%{search_branch}%'] * 3)

    if search_serial:
        query += """
            AND (
                ws.serial_number LIKE ?
                OR ws.firmware_number LIKE ?
            )
        """
        params.extend([
            f'%{search_serial}%',
            f'%{search_serial}%'
        ])

    if  search_app_version:
        query += " AND b.app_version = ?"
        params.append(search_app_version)   
    
    if search_state:
        query += " AND b.state LIKE ?"
        params.append(f'%{search_state}%')
    
    query += " ORDER BY s.name, b.branch_name, ws.serial_number"
    
    results = db.execute(query, params).fetchall()
    
    # Get unique values for filter dropdowns
    supermarkets = db.execute("SELECT DISTINCT name FROM supermarkets ORDER BY name").fetchall()
    states = db.execute("SELECT DISTINCT state FROM branches ORDER BY state").fetchall()
    app_versions = db.execute("SELECT DISTINCT app_version FROM branches WHERE app_version IS NOT NULL AND app_version != '' ORDER BY app_version").fetchall()
        
    db.close()
    
    return render_template('monitoring.html', 
                         results=results,
                         supermarkets=supermarkets,
                         states=states,
                         app_versions=app_versions,
                         search_supermarket=search_supermarket,
                         search_branch=search_branch,
                         search_serial=search_serial,
                         search_state=search_state,
                         search_app_version=search_app_version)

# Supermarket routes
@app.route('/supermarkets')
@login_required
def supermarkets():
    db = get_db()
    supermarkets = db.execute("SELECT * FROM supermarkets ORDER BY name").fetchall()
    db.close()
    return render_template('supermarkets.html', supermarkets=supermarkets)

@app.route('/supermarkets/add', methods=['GET', 'POST'])
@login_required
def add_supermarket():
    if request.method == 'POST':
        db = get_db()
        cursor = db.execute(
            """INSERT INTO supermarkets (name, head_office_address, contact_person, phone, email, remarks)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (request.form['name'], request.form['address'], request.form['contact_person'],
             request.form['phone'], request.form['email'], request.form['remarks'])
        )
        record_id = cursor.lastrowid
        db.commit()
        
        # Log the action - for create, log the name as the new data
        log_action('create', 'supermarkets', record_id, None, request.form['name'])
        
        db.close()
        flash('Supermarket added successfully!', 'success')
        return redirect(url_for('supermarkets'))
    return render_template('supermarket_form.html', supermarket=None)

@app.route('/supermarkets/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_supermarket(id):
    db = get_db()
    if request.method == 'POST':
        # Get old data before updating
        old_data = db.execute("SELECT * FROM supermarkets WHERE id = ?", (id,)).fetchone()
        old_data_dict = dict(old_data) if old_data else None
        
        db.execute(
            """UPDATE supermarkets SET name=?, head_office_address=?, contact_person=?, 
               phone=?, email=?, remarks=? WHERE id=?""",
            (request.form['name'], request.form['address'], request.form['contact_person'],
             request.form['phone'], request.form['email'], request.form['remarks'], id)
        )
        db.commit()
        
        # Log the action
        new_data = {
            'name': request.form['name'],
            'head_office_address': request.form['address'],
            'contact_person': request.form['contact_person'],
            'phone': request.form['phone'],
            'email': request.form['email'],
            'remarks': request.form['remarks']
        }
        log_action('update', 'supermarkets', id, old_data_dict, new_data)
        
        db.close()
        flash('Supermarket updated successfully!', 'success')
        return redirect(url_for('supermarkets'))
    
    supermarket = db.execute("SELECT * FROM supermarkets WHERE id = ?", (id,)).fetchone()
    db.close()
    return render_template('supermarket_form.html', supermarket=supermarket)

@app.route('/supermarkets/delete/<int:id>')
@login_required
@admin_required
def delete_supermarket(id):
    db = get_db()
    # Get data before deleting
    old_data = db.execute("SELECT * FROM supermarkets WHERE id = ?", (id,)).fetchone()
    old_data_dict = dict(old_data) if old_data else None
    
    db.execute("DELETE FROM supermarkets WHERE id = ?", (id,))
    db.commit()
    
    # Log the action - for delete, log the name as the old data
    log_action('delete', 'supermarkets', id, old_data_dict['name'] if old_data_dict else None, None)
    
    db.close()
    flash('Supermarket deleted successfully!', 'success')
    return redirect(url_for('supermarkets'))

# Branch routes
@app.route('/branches')
@login_required
def branches():
    db = get_db()

    # Get filter parameters from query string
    search_query = request.args.get('search', '').strip().lower()
    supermarket_filter = request.args.get('supermarket', '').strip()
    state_filter = request.args.get('state', '').strip()

    # Base query
    query = """
        SELECT b.*, s.name AS supermarket_name
        FROM branches b
        JOIN supermarkets s ON b.supermarket_id = s.id
        WHERE 1=1
    """
    params = []

    # Add filters dynamically
    if search_query:
        query += """ AND (
            LOWER(b.branch_name) LIKE ? OR
            LOWER(b.branch_code) LIKE ? OR
            LOWER(s.name) LIKE ? OR
            LOWER(b.state) LIKE ? OR
            LOWER(b.app_version) LIKE ? OR
            LOWER(b.address) LIKE ?
        )"""
        like_value = f"%{search_query}%"
        params.extend([like_value] * 6)

    if supermarket_filter:
        query += " AND b.supermarket_id = ?"
        params.append(supermarket_filter)

    if state_filter:
        query += " AND LOWER(b.state) = ?"
        params.append(state_filter.lower())

    query += " ORDER BY s.name, b.branch_name"

    branches = db.execute(query, params).fetchall()

    # Dropdown filter options
    supermarkets = db.execute("SELECT id, name FROM supermarkets ORDER BY name").fetchall()
    states = db.execute("SELECT DISTINCT state FROM branches ORDER BY state").fetchall()

    db.close()

    return render_template(
        'branches.html',
        branches=branches,
        supermarkets=supermarkets,
        states=states,
        filters={
            'search': search_query,
            'supermarket': supermarket_filter,
            'state': state_filter
        }
    )

@app.route('/branches/check_code')
@login_required
def check_branch_code():
    code = request.args.get('branch_code', '').strip()
    branch_id = request.args.get('branch_id', type=int, default=0)  # for edit
    db = get_db()
    exists = db.execute(
        "SELECT 1 FROM branches WHERE LOWER(branch_code) = LOWER(?) AND id != ?",
        (code, branch_id)
    ).fetchone()
    db.close()
    return {'exists': bool(exists)}

@app.route('/branches/add', methods=['GET', 'POST'])
@login_required
def add_branch():
    db = get_db()
    
    if request.method == 'POST':
        branch_code = request.form['branch_code'].strip()
        
        # Check if branch_code already exists
        if branch_code:
            existing = db.execute(
                "SELECT 1 FROM branches WHERE branch_code = ?", (branch_code,)
            ).fetchone()
            if existing:
                flash(f"Branch code '{branch_code}' already exists!", 'danger')
                supermarkets = db.execute("SELECT id, name FROM supermarkets ORDER BY name").fetchall()
                db.close()
                return render_template('branch_form.html', branch=None, supermarkets=supermarkets)
        
        # Insert new branch
        cursor = db.execute(
            """INSERT INTO branches (supermarket_id, branch_name, branch_code, branch_region, state, app_version, address, contact_person, 
               phone, branch_total, opening_date, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                request.form['supermarket_id'],
                request.form['branch_name'],
                branch_code,
                request.form['branch_region'],
                request.form['state'],
                request.form['app_version'],
                request.form['address'],
                request.form['contact_person'],
                request.form['phone'],
                request.form['branch_total'],
                request.form['opening_date'],
                request.form['notes']
            )
        )
        record_id = cursor.lastrowid
        db.commit()
        
        # Log the action
        log_action('create', 'branches', record_id, None, request.form['branch_name'])
        
        db.close()
        flash('Branch added successfully!', 'success')
        return redirect(url_for('branches'))
    
    # GET request
    supermarkets = db.execute("SELECT id, name FROM supermarkets ORDER BY name").fetchall()
    db.close()
    return render_template('branch_form.html', branch=None, supermarkets=supermarkets)

@app.route('/branches/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_branch(id):
    db = get_db()
    
    if request.method == 'POST':
        # Get old data
        old_data = db.execute("SELECT * FROM branches WHERE id = ?", (id,)).fetchone()
        old_data_dict = dict(old_data) if old_data else {}

        # Convert datetime objects to strings
        for k, v in old_data_dict.items():
            if isinstance(v, (datetime.datetime, date)):
                old_data_dict[k] = v.strftime('%Y-%m-%d')

        # Prepare new data
        new_data = {
            'supermarket_id': request.form['supermarket_id'],
            'branch_name': request.form['branch_name'],
            'branch_code': request.form['branch_code'].strip(),
            'branch_region': request.form['branch_region'],
            'state': request.form['state'],
            'app_version': request.form['app_version'],
            'address': request.form['address'],
            'contact_person': request.form['contact_person'],
            'phone': request.form['phone'],
            'branch_total': request.form['branch_total'],
            'opening_date': request.form['opening_date'],
            'notes': request.form['notes']
        }

        # Check if branch_code is changing and already exists on another branch
        if new_data['branch_code'] and new_data['branch_code'] != old_data_dict.get('branch_code'):
            existing = db.execute(
                "SELECT 1 FROM branches WHERE branch_code = ? AND id != ?",
                (new_data['branch_code'], id)
            ).fetchone()
            if existing:
                flash(f"Branch code '{new_data['branch_code']}' already exists!", 'danger')
                supermarkets = db.execute("SELECT id, name FROM supermarkets ORDER BY name").fetchall()
                db.close()
                return render_template('branch_form.html', branch=old_data, supermarkets=supermarkets)

        # Calculate only changed fields
        old_change, new_change = get_changes(old_data_dict, new_data)

        # Update DB
        db.execute(
            """UPDATE branches SET supermarket_id=?, branch_name=?, branch_code=?, branch_region=?, state=?, app_version=?, address=?, 
               contact_person=?, phone=?, branch_total=?, opening_date=?, notes=? WHERE id=?""",
            tuple(new_data.values()) + (id,)
        )
        db.commit()

        # Log only changed fields
        if old_change and new_change:
            log_action('update', 'branches', id, old_change, new_change)

        db.close()
        flash('Branch updated successfully!', 'success')
        return redirect(url_for('branches'))

    # GET request
    branch = db.execute("SELECT * FROM branches WHERE id = ?", (id,)).fetchone()
    supermarkets = db.execute("SELECT id, name FROM supermarkets ORDER BY name").fetchall()
    db.close()
    return render_template('branch_form.html', branch=branch, supermarkets=supermarkets)


@app.route('/branches/delete/<int:id>')
@login_required
@admin_required
def delete_branch(id):
    db = get_db()
    # Get data before deleting
    old_data = db.execute("SELECT * FROM branches WHERE id = ?", (id,)).fetchone()
    old_data_dict = dict(old_data) if old_data else None
    
    db.execute("DELETE FROM branches WHERE id = ?", (id,))
    db.commit()
    
    # Log the action - for delete, log the branch name as the old data
    log_action('delete', 'branches', id, old_data_dict['branch_name'] if old_data_dict else None, None)
    
    db.close()
    flash('Branch deleted successfully!', 'success')
    return redirect(url_for('branches'))

# API endpoint for getting branches by supermarket - FIXED VERSION
@app.route('/api/branches/<int:supermarket_id>')
@login_required
def get_branches_by_supermarket(supermarket_id):
    db = get_db()
    branches = db.execute(
        "SELECT id, branch_name as name, branch_code as code FROM branches WHERE supermarket_id = ? ORDER BY branch_name",
        (supermarket_id,)
    ).fetchall()
    db.close()
    return jsonify([dict(b) for b in branches])
    
# returns branches (optionally filtered by supermarket_id and/or state) including state for each branch
@app.route('/get_branches')
@login_required
def get_branches():
    supermarket_id = request.args.get('supermarket_id', '').strip()
    state = request.args.get('state', '').strip()

    db = get_db()
    query = "SELECT id, branch_name, branch_code, state FROM branches WHERE 1=1"
    params = []
    if supermarket_id:
        query += " AND supermarket_id = ?"
        params.append(supermarket_id)
    if state:
        query += " AND state = ?"
        params.append(state)
    query += " ORDER BY branch_name"
    rows = db.execute(query, params).fetchall()
    db.close()

    result = []
    for r in rows:
        result.append({
            "id": r["id"],
            "branch_name": r["branch_name"],
            "branch_code": r["branch_code"],
            "state": r["state"],
            # human-friendly display text used in dropdown
            "name": f"{r['branch_name']} | {r['branch_code']}"
        })
    return jsonify(result)


# returns distinct states and branches for a supermarket (supermarket_id is required here)
@app.route('/get_states_branches/<int:supermarket_id>')
@login_required
def get_states_branches(supermarket_id):
    db = get_db()

    states_rows = db.execute(
        "SELECT DISTINCT state FROM branches WHERE supermarket_id = ? ORDER BY state",
        (supermarket_id,)
    ).fetchall()
    branches_rows = db.execute(
        "SELECT id, branch_name, branch_code, state FROM branches WHERE supermarket_id = ? ORDER BY branch_name",
        (supermarket_id,)
    ).fetchall()

    db.close()

    return jsonify({
        "states": [r["state"] for r in states_rows],
        "branches": [
            {"id": r["id"], "branch_name": r["branch_name"], "branch_code": r["branch_code"], "state": r["state"],
             "name": f"{r['branch_name']} | {r['branch_code']}"}
            for r in branches_rows
        ]
    })

    
@app.route('/scales')
@login_required
def scales():
    db = get_db()
    
    # Get filter parameters from query string
    supermarket_filter = request.args.get('supermarket', '')
    branch_filter = request.args.get('branch', '')
    state_filter = request.args.get('state', '')
    status_filter = request.args.get('status', '')
    
    query = """
        SELECT ws.*, s.name as supermarket, b.branch_name, b.branch_code, b.state
        FROM weighing_scales ws
        JOIN supermarkets s ON ws.supermarket_id = s.id
        JOIN branches b ON ws.branch_id = b.id
        WHERE 1=1
    """
    params = []
    
    if supermarket_filter:
        query += " AND ws.supermarket_id = ?"
        params.append(supermarket_filter)
    if branch_filter:
        query += " AND ws.branch_id = ?"
        params.append(branch_filter)
    if state_filter:
        query += " AND b.state = ?"
        params.append(state_filter)
    if status_filter:
        query += " AND ws.maintenance_status = ?"
        params.append(status_filter)
    
    query += " ORDER BY s.name, b.branch_name, b.branch_code"
    scales = db.execute(query, params).fetchall()
    
     # 🧮 Count total number of scales
    total_scales = len(scales)

    # (Optional) Count per branch
    branch_counts = db.execute("""
        SELECT b.branch_name, COUNT(*) AS total
        FROM weighing_scales ws
        JOIN branches b ON ws.branch_id = b.id
        GROUP BY b.branch_name
        ORDER BY b.branch_name
    """).fetchall()
    
    # Load dropdown options
    supermarkets = db.execute("SELECT id, name FROM supermarkets ORDER BY name").fetchall()
    states = db.execute("SELECT DISTINCT state FROM branches ORDER BY state").fetchall()
    branches = db.execute("SELECT id, branch_name, branch_code FROM branches ORDER BY branch_name").fetchall()
    
    db.close()
    
    return render_template(
        'scales.html',
        scales=scales,
        supermarkets=supermarkets,
        states=states,
        branches=branches,
        filters={
            'supermarket': supermarket_filter,
            'branch': branch_filter,
            'state': state_filter,
            'status': status_filter
        }
    )

@app.route('/scales/add', methods=['GET', 'POST'])
@login_required
def add_scale():
    db = get_db()
    try:
        if request.method == 'POST':
            document_path = None

            # Handle file upload
            if 'document' in request.files:
                file = request.files['document']
                if file and file.filename and allowed_file(file.filename):
                    filename = secure_filename(f"{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    document_path = filename

            # Check for duplicate serial number
            existing = db.execute(
                "SELECT id FROM weighing_scales WHERE serial_number = ?",
                (request.form['serial_number'],)
            ).fetchone()

            if existing:
                flash('Error: Serial number already exists.', 'danger')
                db.close()
                return redirect(url_for('add_scale'))

            # Get form values with defaults
            wintec_sdk = request.form.get('wintec_sdk', '')
            printer_firmware = request.form.get('printer_firmware_number', '')
            
            # Insert new record
            cursor = db.execute(
                """INSERT INTO weighing_scales (
                    supermarket_id, branch_id, brand, model, serial_number,
                    firmware_number, printer_firmware_number, installation_date, ip_address, mac_address,
                    anydesk_id, anydesk_password, weight_license_number,
                    license_expiry_date, maintenance_status, technician_in_charge, wintec_sdk,
                    remarks, document_path
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    request.form['supermarket_id'], 
                    request.form['branch_id'],
                    request.form['brand'], 
                    request.form['model'],
                    request.form['serial_number'], 
                    request.form['firmware_number'],
                    printer_firmware,
                    request.form['installation_date'], 
                    request.form['ip_address'],
                    request.form['mac_address'], 
                    request.form['anydesk_id'],
                    request.form['anydesk_password'], 
                    request.form['weight_license_number'],
                    request.form['license_expiry_date'], 
                    request.form['maintenance_status'],
                    request.form['technician_in_charge'], 
                    wintec_sdk, 
                    request.form['remarks'], 
                    document_path
                )
            )

            record_id = cursor.lastrowid
            db.commit()

            # Log the action
            log_action('create', 'weighing_scales', record_id, None, request.form['serial_number'])

            flash('Weighing scale added successfully!', 'success')
            db.close()
            return redirect(url_for('scales'))

        # GET method — load form
        supermarkets = db.execute("SELECT id, name FROM supermarkets ORDER BY name").fetchall()
        db.close()
        return render_template('scale_form.html', scale=None, supermarkets=supermarkets)

    except KeyError as e:
        db.rollback()
        db.close()
        flash(f'Missing required field: {str(e)}', 'danger')
        return redirect(url_for('add_scale'))

    except sqlite3.IntegrityError as e:
        db.rollback()
        db.close()
        flash(f'Database integrity error: {str(e)}', 'danger')
        return redirect(url_for('add_scale'))

    except sqlite3.OperationalError as e:
        db.rollback()
        db.close()
        flash(f'Database operational error: {str(e)}', 'danger')
        return redirect(url_for('add_scale'))

    except Exception as e:
        db.rollback()
        db.close()
        flash(f'Unexpected error: {str(e)}', 'danger')
        traceback.print_exc()  # This will help you see the full error in console
        return redirect(url_for('add_scale'))

@app.route('/scales/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_scale(id):
    db = get_db()
    if request.method == 'POST':
        # Get old data before updating
        old_data = db.execute("SELECT * FROM weighing_scales WHERE id = ?", (id,)).fetchone()
        
        old_data_dict = {}
        if old_data:
            for key in old_data.keys():
                value = old_data[key]
                # Convert datetime and date objects to string
                if hasattr(value, 'strftime'):
                    old_data_dict[key] = value.strftime('%Y-%m-%d')
                else:
                    old_data_dict[key] = value

        # Handle document upload
        document_path = old_data_dict.get('document_path')
        if 'document' in request.files:
            file = request.files['document']
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                document_path = filename

        # Prepare new data from form
        new_data_dict = {
            'supermarket_id': request.form['supermarket_id'],
            'branch_id': request.form['branch_id'],
            'brand': request.form['brand'],
            'model': request.form['model'],
            'serial_number': request.form['serial_number'],
            'firmware_number': request.form['firmware_number'],
            'printer_firmware_number': request.form['printer_firmware_number'],
            'installation_date': request.form['installation_date'],
            'ip_address': request.form['ip_address'],
            'mac_address': request.form['mac_address'],
            'anydesk_id': request.form['anydesk_id'],
            'anydesk_password': request.form['anydesk_password'],
            'weight_license_number': request.form['weight_license_number'],
            'license_expiry_date': request.form['license_expiry_date'],
            'maintenance_status': request.form['maintenance_status'],
            'technician_in_charge': request.form['technician_in_charge'],
            'wintec_sdk': request.form['wintec_sdk'],
            'remarks': request.form['remarks'],
            'document_path': document_path
        }

        # Compare old and new data to log only changes
        changed_old_data = {}
        changed_new_data = {}
        for key, new_value in new_data_dict.items():
            old_value = old_data_dict.get(key)
            # normalize None and empty string
            old_val_norm = '' if old_value is None else str(old_value)
            new_val_norm = '' if new_value is None else str(new_value)
            if old_val_norm != new_val_norm:
                changed_old_data[key] = old_value
                changed_new_data[key] = new_value

        # Update the database
        db.execute(
            """UPDATE weighing_scales SET supermarket_id=?, branch_id=?, brand=?, model=?, serial_number=?,
               firmware_number=?, printer_firmware_number=?, installation_date=?, ip_address=?, mac_address=?, anydesk_id=?, anydesk_password=?,
               weight_license_number=?, license_expiry_date=?, maintenance_status=?, 
               technician_in_charge=?, wintec_sdk=?, remarks=?, document_path=? WHERE id=?""",
            tuple(new_data_dict.values()) + (id,)
        )
        db.commit()

        # Log only if there are changes
        if changed_old_data:
            log_action('update', 'weighing_scales', id, changed_old_data, changed_new_data)
        
        db.close()
        flash('Weighing scale updated successfully!', 'success')
        return redirect(url_for('scales'))

    # GET request
    scale = db.execute("SELECT * FROM weighing_scales WHERE id = ?", (id,)).fetchone()
    supermarkets = db.execute("SELECT id, name FROM supermarkets ORDER BY name").fetchall()
    branches = db.execute(
        "SELECT id, branch_name FROM branches WHERE supermarket_id = ? ORDER BY branch_name",
        (scale['supermarket_id'],)
    ).fetchall()
    db.close()
    return render_template('scale_form.html', scale=scale, supermarkets=supermarkets, branches=branches)


@app.route('/scales/delete/<int:id>')
@login_required
@admin_required
def delete_scale(id):
    db = get_db()
    # Get data before deleting
    old_data = db.execute("SELECT * FROM weighing_scales WHERE id = ?", (id,)).fetchone()
    old_data_dict = dict(old_data) if old_data else None
    
    db.execute("DELETE FROM weighing_scales WHERE id = ?", (id,))
    db.commit()
    
    # Log the action - for delete, log the serial number as the old data
    log_action('delete', 'weighing_scales', id, old_data_dict['serial_number'] if old_data_dict else None, None)
    
    db.close()
    flash('Weighing scale deleted successfully!', 'success')
    return redirect(url_for('scales'))

@app.route('/scales/<int:id>')
@login_required
def scale_detail(id):
    db = get_db()

    # Get the scale details
    scale = db.execute(
        """SELECT ws.*, 
                  s.name as supermarket, 
                  s.contact_person as sm_contact, 
                  s.phone as sm_phone,
                  b.branch_name, 
                  b.state, 
                  b.app_version, 
                  b.address, 
                  b.branch_code, 
                  b.branch_region, 
                  b.contact_person as br_contact, 
                  b.phone as br_phone
           FROM weighing_scales ws
           JOIN supermarkets s ON ws.supermarket_id = s.id
           JOIN branches b ON ws.branch_id = b.id
           WHERE ws.id = ?""",
        (id,)
    ).fetchone()

    if not scale:
        flash("Scale not found.", "danger")
        return redirect(url_for('scales'))

    # Convert to dict
    scale = dict(scale)
    
    # Calculate expiry status
    if scale['license_expiry_date']:
        from datetime import date as date_class
        
        # Convert to date object if it's a string
        expiry_date = scale['license_expiry_date']
        if isinstance(expiry_date, str):
            expiry_date = datetime.datetime.strptime(expiry_date, '%Y-%m-%d').date()
        
        today = datetime.datetime.now().date()
        days_diff = (expiry_date - today).days
        
        # Add expiry info to scale dict
        scale['expiry_days'] = days_diff
        
        if days_diff < 0:
            scale['expiry_status'] = 'expired'
            scale['expiry_badge'] = 'danger'
            scale['expiry_message'] = f"Expired {abs(days_diff)} days ago"
        elif days_diff == 0:
            scale['expiry_status'] = 'today'
            scale['expiry_badge'] = 'danger'
            scale['expiry_message'] = "Expires Today!"
        elif days_diff <= 7:
            scale['expiry_status'] = 'critical'
            scale['expiry_badge'] = 'danger'
            scale['expiry_message'] = f"Expires in {days_diff} days"
        elif days_diff <= 30:
            scale['expiry_status'] = 'warning'
            scale['expiry_badge'] = 'warning'
            scale['expiry_message'] = f"Expires in {days_diff} days"
        elif days_diff <= 90:
            scale['expiry_status'] = 'notice'
            scale['expiry_badge'] = 'info'
            scale['expiry_message'] = f"Expires in {days_diff} days"
        else:
            scale['expiry_status'] = 'active'
            scale['expiry_badge'] = 'success'
            scale['expiry_message'] = f"Active ({days_diff} days)"

    # Count total scales in the same branch
    total_scales = db.execute(
        """SELECT COUNT(*) FROM weighing_scales 
           WHERE branch_id = ?""",
        (scale['branch_id'],)
    ).fetchone()[0]

    # Maintenance records
    maintenance_history = db.execute(
        """SELECT * FROM maintenance_records 
           WHERE scale_id = ? 
           ORDER BY service_date DESC""",
        (id,)
    ).fetchall()

    db.close()

    return render_template(
        'scale_detail.html',
        scale=scale,
        maintenance_history=maintenance_history,
        total_scales=total_scales
    )
# Maintenance routes
@app.route('/maintenance')
@login_required
def maintenance():
    db = get_db()
    records = db.execute(
        """SELECT mr.*, ws.serial_number, ws.model, s.name as supermarket, b.branch_name
           FROM maintenance_records mr
           JOIN weighing_scales ws ON mr.scale_id = ws.id
           JOIN supermarkets s ON ws.supermarket_id = s.id
           JOIN branches b ON ws.branch_id = b.id
           ORDER BY mr.service_date DESC"""
    ).fetchall()
    db.close()
    return render_template('maintenance.html', records=records)

@app.route('/maintenance/add/<int:scale_id>', methods=['GET', 'POST'])
@login_required
def add_maintenance(scale_id):
    db = get_db()
    if request.method == 'POST':
        cursor = db.execute(
            """INSERT INTO maintenance_records (scale_id, technician_name, service_date,
               issue_description, resolution, next_service_due)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (scale_id, request.form['technician_name'], request.form['service_date'],
             request.form['issue_description'], request.form['resolution'],
             request.form['next_service_due'])
        )
        record_id = cursor.lastrowid
        db.commit()
        
        # Log the action - for create, log the technician name and service date
        log_action('create', 'maintenance_records', record_id, None, 
                  f"{request.form['technician_name']} - {request.form['service_date']}")
        
        db.close()
        flash('Maintenance record added successfully!', 'success')
        return redirect(url_for('scale_detail', id=scale_id))
    
    scale = db.execute(
        """SELECT ws.*, s.name as supermarket, b.branch_name
           FROM weighing_scales ws
           JOIN supermarkets s ON ws.supermarket_id = s.id
           JOIN branches b ON ws.branch_id = b.id
           WHERE ws.id = ?""",
        (scale_id,)
    ).fetchone()
    db.close()
    return render_template('maintenance_form.html', record=None, scale=scale)

@app.route('/export/scales/excel')
@login_required
def export_scales_excel():
    db = get_db()

    # --- Get filters from query string ---
    search_query = request.args.get('search', '').strip().lower()
    supermarket_filter = request.args.get('supermarket', '').strip()
    state_filter = request.args.get('state', '').strip()
    branch_filter = request.args.get('branch', '').strip()

    # --- Base SQL query ---
    query = """
        SELECT ws.serial_number, ws.model, s.name AS supermarket, b.branch_name, b.state,
               ws.installation_date, ws.ip_address, ws.weight_license_number,
               ws.license_expiry_date, ws.maintenance_status, ws.technician_in_charge
        FROM weighing_scales ws
        JOIN supermarkets s ON ws.supermarket_id = s.id
        JOIN branches b ON ws.branch_id = b.id
        WHERE 1=1
    """
    params = []

    # --- Apply search filter (matches multiple columns) ---
    if search_query:
        query += """ AND (
            LOWER(ws.serial_number) LIKE ? OR
            LOWER(ws.model) LIKE ? OR
            LOWER(s.name) LIKE ? OR
            LOWER(b.branch_name) LIKE ? OR
            LOWER(b.state) LIKE ? OR
            LOWER(ws.ip_address) LIKE ? OR
            LOWER(ws.technician_in_charge) LIKE ?
        )"""
        like_value = f"%{search_query}%"
        params.extend([like_value] * 7)

    # --- Apply dropdown filters ---
    if supermarket_filter:
        query += " AND ws.supermarket_id = ?"
        params.append(supermarket_filter)

    if state_filter:
        query += " AND LOWER(b.state) = ?"
        params.append(state_filter.lower())

    if branch_filter:
        query += " AND ws.branch_id = ?"
        params.append(branch_filter)

    # --- Final ordering ---
    query += " ORDER BY s.name, b.branch_name"

    # --- Execute query ---
    scales = db.execute(query, params).fetchall()
    db.close()

    # --- Log export action ---
    log_action('export', 'weighing_scales', None, None, {'format': 'excel'})

    # --- Create Excel workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Weighing Scales"

    # Headers
    headers = [
        'Serial Number', 'Model', 'Supermarket', 'Branch', 'State',
        'Installation Date', 'IP Address', 'License Number', 'License Expiry',
        'Status', 'Technician'
    ]
    ws.append(headers)

    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for scale in scales:
        ws.append([
            scale['serial_number'],
            scale['model'],
            scale['supermarket'],
            scale['branch_name'],
            scale['state'],
            scale['installation_date'],
            scale['ip_address'],
            scale['weight_license_number'],
            scale['license_expiry_date'],
            scale['maintenance_status'],
            scale['technician_in_charge'],
        ])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # --- Send Excel file ---
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'weighing_scales_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx'
    )

@app.route('/export/scales/pdf')
@login_required
def export_scales_pdf():
    db = get_db()

    # Get filters from query string (same as scales route)
    supermarket_filter = request.args.get('supermarket', '').strip()
    state_filter = request.args.get('state', '').strip()
    branch_filter = request.args.get('branch', '').strip()

    # Base query
    query = """
        SELECT ws.serial_number, ws.model, s.name AS supermarket, 
               b.branch_name, b.branch_code, b.state,
               ws.license_expiry_date, ws.maintenance_status
        FROM weighing_scales ws
        JOIN supermarkets s ON ws.supermarket_id = s.id
        JOIN branches b ON ws.branch_id = b.id
        WHERE 1=1
    """
    params = []

    # Apply filters (same logic as your main view)
    if supermarket_filter:
        query += " AND ws.supermarket_id = ?"
        params.append(supermarket_filter)

    if state_filter:
        query += " AND LOWER(b.state) = ?"
        params.append(state_filter.lower())

    if branch_filter:
        query += " AND ws.branch_id = ?"
        params.append(branch_filter)

    query += " ORDER BY s.name, b.branch_name"

    scales = db.execute(query, params).fetchall()
    db.close()

    log_action('export', 'weighing_scales', None, None, {'format': 'pdf', 'filters': request.args.to_dict()})

    # PDF generation
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2C3E50'),
        spaceAfter=30,
        alignment=1
    )

    title = Paragraph(f"Weighing Scales Report - {datetime.datetime.now().strftime('%B %d, %Y')}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))

    data = [['No.', 'Serial No.', 'Model', 'Supermarket', 'Branch', 'Branch Code', 'State', 'Status']]
    for index, scale in enumerate(scales, start=1):
        data.append([
            index,  # Running number
            scale['serial_number'],
            scale['model'],
            scale['supermarket'][:20],
            scale['branch_name'][:15],
            scale['branch_code'] or '',
            scale['state'],
            #scale['license_expiry_date'] or 'N/A',
            scale['maintenance_status']
        ])

    table = Table(
        data, 
        colWidths=[0.5*inch, 1*inch, 1*inch, 1.3*inch, 1.3*inch, 0.9*inch, 0.8*inch, 1*inch, 0.8*inch]
    )
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'weighing_scales_{datetime.datetime.now().strftime("%Y%m%d")}.pdf'
    )

# User Management routes
@app.route('/users')
@login_required
@admin_required
def users():
    if current_user.role != 'admin':
        flash('Unauthorized access', 'danger')
        return redirect(url_for('dashboard'))
    
    db = get_db()
    users = db.execute("""
        SELECT id, username, full_name, email, role, is_active, created_at 
        FROM users 
        ORDER BY created_at DESC
    """).fetchall()
    db.close()
    
    return render_template('users.html', users=users)

@app.route('/users/<int:id>/toggle_status', methods=['POST'])
@login_required
def toggle_user_status(id):
    if current_user.role != 'admin':
        flash('Unauthorized', 'error')
        return redirect(url_for('users'))
    
    # Prevent admin from disabling themselves
    if id == current_user.id:
        flash('You cannot disable your own account', 'error')
        return redirect(url_for('users'))
    
    db = get_db()
    
    # Get current user status
    user = db.execute("SELECT id, username, is_active FROM users WHERE id = ?", (id,)).fetchone()
    
    if not user:
        db.close()
        flash('User not found', 'error')
        return redirect(url_for('users'))
    
    # Toggle the is_active status
    new_status = 0 if user['is_active'] else 1
    db.execute("UPDATE users SET is_active = ? WHERE id = ?", (new_status, id))
    db.commit()
    
    # Log the action
    status_text = 'enabled' if new_status else 'disabled'
    log_action('toggle_status', 'users', id, None, f"User {user['username']} {status_text}")
    
    db.close()
    
    flash(f'User {user["username"]} has been {status_text}', 'success')
    return redirect(url_for('users'))

@app.route('/users/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_user():
    if request.method == 'POST':
        db = get_db()
        hashed_password = generate_password_hash(request.form['password'])
        try:
            cursor = db.execute(
                """INSERT INTO users (username, password, full_name, email, role)
                   VALUES (?, ?, ?, ?, ?)""",
                (request.form['username'], hashed_password, request.form['full_name'],
                 request.form['email'], request.form['role'])
            )
            record_id = cursor.lastrowid
            db.commit()
            
            # Log the action - for create, log the username as the new data
            log_action('create', 'users', record_id, None, request.form['username'])
            
            flash('User added successfully!', 'success')
        except sqlite3.IntegrityError:
            flash('Username already exists!', 'danger')
        db.close()
        return redirect(url_for('users'))
    return render_template('user_form.html', user=None)

@app.route('/users/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(id):
    db = get_db()
    if request.method == 'POST':
        # Get old data before updating
        old_data = db.execute("SELECT * FROM users WHERE id = ?", (id,)).fetchone()
        old_data_dict = dict(old_data) if old_data else None
        
        if request.form.get('password'):
            hashed_password = generate_password_hash(request.form['password'])
            db.execute(
                """UPDATE users SET username=?, password=?, full_name=?, email=?, role=?
                   WHERE id=?""",
                (request.form['username'], hashed_password, request.form['full_name'],
                 request.form['email'], request.form['role'], id)
            )
        else:
            db.execute(
                """UPDATE users SET username=?, full_name=?, email=?, role=? WHERE id=?""",
                (request.form['username'], request.form['full_name'], request.form['email'],
                 request.form['role'], id)
            )
        db.commit()
        
        # Log the action
        new_data = {
            'username': request.form['username'],
            'full_name': request.form['full_name'],
            'email': request.form['email'],
            'role': request.form['role']
        }
        log_action('update', 'users', id, old_data_dict, new_data)
        
        db.close()
        flash('User updated successfully!', 'success')
        return redirect(url_for('users'))
    
    user = db.execute("SELECT * FROM users WHERE id = ?", (id,)).fetchone()
    db.close()
    return render_template('user_form.html', user=user)

@app.route('/users/delete/<int:id>')
@login_required
@admin_required
def delete_user(id):
    if id == current_user.id:
        flash('You cannot delete your own account!', 'danger')
        return redirect(url_for('users'))
    
    db = get_db()
    # Get data before deleting
    old_data = db.execute("SELECT * FROM users WHERE id = ?", (id,)).fetchone()
    old_data_dict = dict(old_data) if old_data else None
    
    db.execute("DELETE FROM users WHERE id = ?", (id,))
    db.commit()
    
    # Log the action - for delete, log the username as the old data
    log_action('delete', 'users', id, old_data_dict['username'] if old_data_dict else None, None)
    
    db.close()
    flash('User deleted successfully!', 'success')
    return redirect(url_for('users'))

# Notifications
@app.route('/notifications/mark_read/<int:id>')
@login_required
def mark_notification_read(id):
    db = get_db()
    # Get old data before updating
    old_data = db.execute("SELECT * FROM notifications WHERE id = ?", (id,)).fetchone()
    old_data_dict = dict(old_data) if old_data else None
    
    db.execute("UPDATE notifications SET is_read = 1 WHERE id = ?", (id,))
    db.commit()
    
    # Log the action
    new_data = {'is_read': 1}
    log_action('update', 'notifications', id, old_data_dict, new_data)
    
    db.close()
    return jsonify({'success': True})

# Settings/Profile
@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    db = get_db()
    if request.method == 'POST':
        # Get old data before updating
        old_data = db.execute("SELECT * FROM users WHERE id = ?", (current_user.id,)).fetchone()
        old_data_dict = dict(old_data) if old_data else None
        
        if request.form.get('current_password'):
            user = db.execute("SELECT * FROM users WHERE id = ?", (current_user.id,)).fetchone()
            if check_password_hash(user['password'], request.form['current_password']):
                new_password = generate_password_hash(request.form['new_password'])
                db.execute("UPDATE users SET password = ? WHERE id = ?", (new_password, current_user.id))
                db.commit()
                
                # Log the action
                new_data = {'password': '[CHANGED]'}
                log_action('update', 'users', current_user.id, old_data_dict, new_data)
                
                flash('Password updated successfully!', 'success')
            else:
                flash('Current password is incorrect!', 'danger')
        else:
            db.execute(
                "UPDATE users SET full_name = ?, email = ? WHERE id = ?",
                (request.form['full_name'], request.form['email'], current_user.id)
            )
            db.commit()
            
            # Log the action
            new_data = {
                'full_name': request.form['full_name'],
                'email': request.form['email']
            }
            log_action('update', 'users', current_user.id, old_data_dict, new_data)
            
            flash('Profile updated successfully!', 'success')
        db.close()
        return redirect(url_for('profile'))
    
    user = db.execute("SELECT * FROM users WHERE id = ?", (current_user.id,)).fetchone()
    db.close()
    return render_template('profile.html', user=user)

@app.route('/logs')
@login_required
def logs():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('dashboard'))
    
    db = get_db()
    
    # Get filter parameters (default to 'today')
    date_filter = request.args.get('date_filter', 'today')
    custom_date = request.args.get('custom_date', '')
    
    # Build query based on filter
    if date_filter == 'all':
        query = 'SELECT * FROM logs ORDER BY timestamp DESC'
        logs_data = db.execute(query).fetchall()
    elif date_filter == 'custom' and custom_date:
        query = '''
            SELECT * FROM logs 
            WHERE DATE(timestamp) = ? 
            ORDER BY timestamp DESC
        '''
        logs_data = db.execute(query, (custom_date,)).fetchall()
    else:  # default to 'today'
        query = '''
            SELECT * FROM logs 
            WHERE DATE(timestamp) = DATE('now') 
            ORDER BY timestamp DESC
        '''
        logs_data = db.execute(query).fetchall()
    
    # Mark all unread logs as read
    db.execute("UPDATE logs SET is_read = 1 WHERE is_read = 0")
    db.commit()
    
    return render_template('logs.html', 
                         logs=logs_data, 
                         date_filter=date_filter,
                         custom_date=custom_date)

@app.route('/logs/clear', methods=['POST'])
@login_required
def clear_logs():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('dashboard'))

    db = get_db()
    db.execute('DELETE FROM logs')
    db.commit()
    #log_action('clear_logs', 'logs', None, None, 'All logs cleared')
    flash('All logs have been cleared successfully.', 'success')
    return redirect(url_for('logs'))
    
@app.route('/check_serial/<serial_number>')
@login_required
def check_serial(serial_number):
    db = get_db()
    existing = db.execute(
        "SELECT ws.id, ws.serial_number, b.branch_name, s.name AS supermarket "
        "FROM weighing_scales ws "
        "LEFT JOIN branches b ON ws.branch_id = b.id "
        "LEFT JOIN supermarkets s ON ws.supermarket_id = s.id "
        "WHERE ws.serial_number = ?", 
        (serial_number.strip(),)
    ).fetchone()
    db.close()

    if existing:
        return jsonify({
            "exists": True,
            "serial_number": existing["serial_number"],
            "branch": existing["branch_name"],
            "supermarket": existing["supermarket"]
        })
    else:
        return jsonify({"exists": False})

@app.route('/branch/<int:id>/scales')
@login_required
def branch_scales(id):
    db = get_db()
    branch = db.execute(
        "SELECT branch_name, app_version FROM branches WHERE id = ?", 
        (id,)
    ).fetchone()

    scales = db.execute(
        """SELECT id, serial_number, model, mac_address, ip_address, installation_date, 
                  firmware_number, maintenance_status
           FROM weighing_scales 
           WHERE branch_id = ?""",
        (id,)
    ).fetchall()
    
    # Convert rows to dict
    scale_list = [dict(row) for row in scales]

    # Add app_version into each scale (if needed)
    for scale in scale_list:
        scale["app_version"] = branch["app_version"] if branch else "-"

    return jsonify({
        "branch_name": branch["branch_name"] if branch else "",
        "app_version": branch["app_version"] if branch else "",
        "scales": scale_list
    })


@app.route('/export/scales/excel/<branch_id>')
@login_required
def export_scales_excel_branch(branch_id):
    import pandas as pd
    from io import BytesIO
    from flask import send_file
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment

    db = get_db()

    # ✅ Get branch + supermarket info in one query
    branch = db.execute("""
        SELECT b.branch_name, b.app_version, s.name AS supermarket_name
        FROM branches b
        LEFT JOIN supermarkets s ON b.supermarket_id = s.id
        WHERE b.id = ?
    """, (branch_id,)).fetchone()

    if not branch:
        flash("Branch not found.", "danger")
        return redirect(url_for('branches'))

    branch_name = branch['branch_name']
    app_version = branch['app_version']
    supermarket_name = branch['supermarket_name'] or "Unknown Supermarket"

    # ✅ Get all scales for this branch
    scales = db.execute("""
        SELECT 
            ws.serial_number, 
            ws.model, 
            ws.ip_address,
            ws.firmware_number,
            ? AS app_version
        FROM weighing_scales ws
        WHERE ws.branch_id = ?
    """, (app_version, branch_id)).fetchall()

    # ✅ Convert to DataFrame
    df = pd.DataFrame(scales, columns=[
        "Serial Number",
        "Model",
        "IP Address",
        "Firmware Number",
        "App Version"
    ])

    # ✅ Prepare Excel in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Start writing DataFrame from row 4 (leave space for title)
        df.to_excel(writer, index=False, sheet_name='Scales', startrow=3)
        writer.close()

    # ✅ Load workbook to add custom title formatting
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    # ✅ Add title text (merged + centered)
    title = f"Scales Report for {supermarket_name} - {branch_name}"
    ws.merge_cells('A1:E1')
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Optional subtitle (App Version)
    ws.merge_cells('A2:E2')
    ws['A2'] = f"App Version: {app_version or '-'}"
    ws['A2'].alignment = Alignment(horizontal='center')

    # ✅ Save workbook back into memory
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    # ✅ Dynamic filename
    safe_branch_name = branch_name.replace(" ", "_").replace("/", "-")
    filename = f"Scales_{safe_branch_name}.xlsx"

    return send_file(
        final_output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/export/scales/pdf/<branch_id>')
@login_required
def export_scales_pdf_branch(branch_id):
    from io import BytesIO
    from flask import send_file
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet

    db = get_db()
    
    # ✅ Get branch + supermarket info in one query
    branch = db.execute("""
        SELECT b.branch_name, b.app_version, s.name AS supermarket_name
        FROM branches b
        LEFT JOIN supermarkets s ON b.supermarket_id = s.id
        WHERE b.id = ?
    """, (branch_id,)).fetchone()

    if not branch:
        flash("Branch not found.", "danger")
        return redirect(url_for('branches'))

    branch_name = branch['branch_name']
    app_version = branch['app_version']
    supermarket_name = branch['supermarket_name'] or "Unknown Supermarket"

    # ✅ Get all scales for this branch
    scales = db.execute("""
        SELECT 
            ws.serial_number, 
            ws.model, 
            ws.ip_address,
            ws.firmware_number,
            ? AS app_version
        FROM weighing_scales ws
        WHERE ws.branch_id = ?
    """, (app_version, branch_id,)).fetchall()

    # ✅ Prepare PDF
    buffer = BytesIO()
    pdf = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    styles = getSampleStyleSheet()
    title = Paragraph(f"<b>Scales Report for {supermarket_name} - {branch_name}</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))

    if not scales:
        elements.append(Paragraph("No scales found for this branch.", styles['Normal']))
    else:
        data = [
            ["No", "Serial Number", "Model", "IP Address", "Firmware Number", "App Version"]
        ]
        for i, s in enumerate(scales, start=1):
            data.append([
                str(i),
                s["serial_number"] or "-",
                s["model"] or "-",
                s["ip_address"] or "-",
                s["firmware_number"] or "-",
                s["app_version"] or "-"
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(table)

    pdf.build(elements)
    buffer.seek(0)

    safe_branch_name = branch_name.replace(" ", "_").replace("/", "-")
    filename = f"Scales_{safe_branch_name}.pdf"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf"
    )
    
@app.route('/api/scales/count')
@login_required
def get_scales_count():
    supermarket_id = request.args.get('supermarket_id')
    db = get_db()
    
    if supermarket_id == 'all':
        count = db.execute("SELECT COUNT(*) as count FROM weighing_scales").fetchone()['count']
    else:
        count = db.execute(
            "SELECT COUNT(*) as count FROM weighing_scales WHERE supermarket_id = ?",
            (supermarket_id,)
        ).fetchone()['count']
    
    db.close()
    
    return jsonify({'total_scales': count})

@app.route('/logs')
@login_required
def show_logs():
    db = get_db()
    cur = db.cursor()

    # Fetch logs
    cur.execute("SELECT * FROM logs ORDER BY id DESC")
    logs = cur.fetchall()

    # Mark all unread logs as read
    cur.execute("UPDATE logs SET is_read = 1 WHERE is_read = 0")
    db.commit()

    # Render the logs template
    return render_template('logs.html', logs=logs)
   
@app.route('/logs/count')
def logs_count():
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT COUNT(*) FROM logs WHERE is_read = 0")
    count = cur.fetchone()[0]
    return {"count": count}
    
@app.route('/logs/mark_read', methods=['POST'])
def mark_logs_read():
    db = get_db()
    cur = db.cursor()
    cur.execute("UPDATE logs SET is_read = 1 WHERE is_read = 0")
    db.commit()
    return {"status": "ok"}
    
@app.route('/api/branches')
@app.route('/api/branches/<int:supermarket_id>')
def api_branches(supermarket_id=None):
    db = get_db()
    if supermarket_id:
        branches = db.execute('SELECT * FROM branches WHERE supermarket_id = ?', (supermarket_id,)).fetchall()
    else:
        branches = db.execute('SELECT * FROM branches').fetchall()
    return jsonify([dict(row) for row in branches])
    
@app.route('/expiring_licenses')
def expiring_licenses():
    db = get_db()
    db.row_factory = sqlite3.Row
    cursor = db.cursor()
    
    query = """
        SELECT 
            ws.id,
            ws.serial_number,
            s.name AS supermarket_name,
            b.branch_name AS branch_name,
            ws.license_expiry_date,
            (SELECT COUNT(*) FROM renewal_history WHERE scale_id = ws.id) as renewal_count
        FROM weighing_scales ws
        LEFT JOIN supermarkets s ON ws.supermarket_id = s.id
        LEFT JOIN branches b ON ws.branch_id = b.id
        WHERE ws.license_expiry_date IS NOT NULL
          AND julianday(ws.license_expiry_date) - julianday('now') <= 30
          AND julianday(ws.license_expiry_date) - julianday('now') >= 0
        ORDER BY ws.license_expiry_date ASC
    """
    
    cursor.execute(query)
    licenses = [dict(row) for row in cursor.fetchall()]
    
    # Calculate days_left in Python (same as dashboard)
    today = datetime.datetime.now().date()
    for license in licenses:
        if license['license_expiry_date']:
            expiry = license['license_expiry_date']
            if isinstance(expiry, str):
                expiry = datetime.datetime.strptime(expiry, '%Y-%m-%d').date()
            license['days_left'] = (expiry - today).days
        else:
            license['days_left'] = None
    
    cursor.close()
    
    return render_template('expiring_licenses.html', licenses=licenses)

@app.route('/renew_license', methods=['POST'])
@login_required
def renew_license():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('expiring_licenses'))

    scale_id = request.form.get('scale_id')
    renew_date_str = request.form.get('renew_date')

    if not scale_id or not renew_date_str:
        flash('Invalid data', 'danger')
        return redirect(url_for('expiring_licenses'))

    try:
        from datetime import datetime, timedelta

        # Convert string to date
        renew_date = datetime.strptime(renew_date_str, '%Y-%m-%d')

        # Add 1 year
        try:
            new_expiry = renew_date.replace(year=renew_date.year + 1)
        except ValueError:
            # Handle Feb 29
            new_expiry = renew_date + (datetime(renew_date.year + 1, 3, 1) - datetime(renew_date.year, 3, 1))

        db = get_db()

        # Get old expiry date for renewal_history
        old_expiry = db.execute('SELECT license_expiry_date FROM weighing_scales WHERE id = ?', (scale_id,)).fetchone()
        old_expiry_date = old_expiry['license_expiry_date'] if old_expiry else None

        # Update the weighing_scales table
        db.execute(
            'UPDATE weighing_scales SET license_expiry_date = ? WHERE id = ?',
            (new_expiry.date(), scale_id)
        )

        # Insert into renewal_history
        db.execute(
            '''
            INSERT INTO renewal_history (scale_id, old_expiry_date, new_expiry_date, renewed_by, notes)
            VALUES (?, ?, ?, ?, ?)
            ''',
            (scale_id, old_expiry_date, new_expiry.date(), current_user.username, 'Scale renewed via admin panel')
        )

        db.commit()
        flash('License renewed successfully and recorded in renewal history!', 'success')

    except Exception as e:
        flash(f'Error: {e}', 'danger')

    return redirect(url_for('expiring_licenses'))

@app.route('/api/renewal-history/<int:scale_id>')
def get_renewal_history(scale_id):
    db = get_db()
    db.row_factory = sqlite3.Row
    cursor = db.cursor()
    
    try:
        query = """
            SELECT 
                rh.id,
                strftime('%Y-%m-%d %H:%M', rh.renewal_date) as renewal_date,
                rh.old_expiry_date,
                rh.new_expiry_date,
                CAST(JULIANDAY(rh.new_expiry_date) - JULIANDAY(rh.old_expiry_date) AS INTEGER) as extended_days,
                rh.renewed_by,
                rh.notes
            FROM renewal_history rh
            WHERE rh.scale_id = ?
            ORDER BY rh.renewal_date DESC
        """
        
        cursor.execute(query, (scale_id,))
        history = [dict(row) for row in cursor.fetchall()]
        
        return jsonify({
            'success': True,
            'total': len(history),
            'history': history
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
    
    finally:
        cursor.close()
        
@app.route('/all_renewal_history')
@login_required
def all_renewal_history():
    db = get_db()
    db.row_factory = sqlite3.Row
    cursor = db.cursor()

    cursor.execute("""
        SELECT rh.id, rh.scale_id, ws.serial_number, rh.old_expiry_date, rh.new_expiry_date,
               rh.renewed_by, rh.renewal_date, rh.notes
        FROM renewal_history rh
        LEFT JOIN weighing_scales ws ON rh.scale_id = ws.id
        ORDER BY rh.renewal_date DESC
    """)
    history = cursor.fetchall()
    cursor.close()

    return render_template('all_renewal_history.html', history=history)
    
@app.route('/export/monitoring/pdf')
@login_required
def export_monitoring_pdf():
    db = get_db()
    cursor = db.cursor()
    
    # Get filter parameters from URL
    search_supermarket = request.args.get('supermarket', '').strip()
    search_branch = request.args.get('branch', '').strip()
    search_serial = request.args.get('serial', '').strip()
    search_brand = request.args.get('brand', '').strip()
    search_state = request.args.get('state', '').strip()
    
    # Get selected columns
    selected_columns_str = request.args.get('columns', '')
    selected_columns = selected_columns_str.split(',') if selected_columns_str else []
    
    # If no columns selected, use all columns
    if not selected_columns:
        selected_columns = ['supermarket', 'branch', 'code', 'state', 'serial', 'firmware', 'ip', 'anydesk', 'version']
    
    # Build SQL query with filters
    query = """
        SELECT 
            s.id,
            s.serial_number,
            s.model,
            s.firmware_number,
            s.ip_address,
            s.anydesk_id,
            b.app_version,
            b.branch_name,
            b.branch_code,
            b.state,
            sm.name as supermarket_name
        FROM weighing_scales s
        JOIN branches b ON s.branch_id = b.id
        JOIN supermarkets sm ON b.supermarket_id = sm.id
        WHERE 1=1
    """
    
    params = []
    
    # Apply filters
    if search_supermarket:
        query += " AND sm.name LIKE ?"
        params.append(f'%{search_supermarket}%')
    
    if search_branch:
        query += " AND (b.branch_name LIKE ? OR b.branch_code LIKE ?)"
        params.append(f'%{search_branch}%')
        params.append(f'%{search_branch}%')
    
    if search_serial:
        query += " AND s.serial_number LIKE ?"
        params.append(f'%{search_serial}%')
    
    if search_brand:
        query += " AND s.model LIKE ?"
        params.append(f'%{search_brand}%')
    
    if search_state:
        query += " AND b.state LIKE ?"
        params.append(f'%{search_state}%')
    
    query += " ORDER BY sm.name, b.branch_name"
    
    # Execute query
    cursor.execute(query, params)
    results = cursor.fetchall()
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4),
        topMargin=0.4*inch, 
        bottomMargin=0.4*inch,
        leftMargin=0.3*inch, 
        rightMargin=0.3*inch
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=8,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    title = Paragraph("Weighing Scales Monitoring Report", title_style)
    elements.append(title)
    
    # Filter info
    filter_info = []
    if search_supermarket:
        filter_info.append(f"Supermarket: {search_supermarket}")
    if search_branch:
        filter_info.append(f"Branch: {search_branch}")
    if search_serial:
        filter_info.append(f"Serial: {search_serial}")
    if search_brand:
        filter_info.append(f"Brand: {search_brand}")
    if search_state:
        filter_info.append(f"State: {search_state}")
    
    if filter_info:
        filter_style = ParagraphStyle(
            'FilterStyle',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#666666')
        )
        filter_text = "Filters: " + " | ".join(filter_info)
        filter_para = Paragraph(filter_text, filter_style)
        elements.append(filter_para)
    
    # Date and count
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#666666')
    )
    date_text = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Records: {len(results)}"
    date_para = Paragraph(date_text, date_style)
    elements.append(date_para)
    elements.append(Spacer(1, 0.15*inch))
    
    # Column configuration
    column_config = {
        'supermarket': {'header': 'Supermarket', 'width': 1.1, 'key': 'supermarket_name', 'wrap': True},
        'branch': {'header': 'Branch', 'width': 1.3, 'key': 'branch_name', 'wrap': True},
        'code': {'header': 'Code', 'width': 0.6, 'key': 'branch_code', 'wrap': False},
        'state': {'header': 'State', 'width': 0.65, 'key': 'state', 'wrap': False},
        'serial': {'header': 'S/No', 'width': 1.2, 'key': 'serial_number', 'wrap': True},
        'firmware': {'header': 'Firmware', 'width': 0.75, 'key': 'firmware_number', 'wrap': False},
        'ip': {'header': 'IP Address', 'width': 1.0, 'key': 'ip_address', 'wrap': True},
        'anydesk': {'header': 'AnyDesk', 'width': 0.8, 'key': 'anydesk_id', 'wrap': False},
        'version': {'header': 'Version', 'width': 0.65, 'key': 'app_version', 'wrap': False}
    }
    
    # Build header and widths based on selected columns
    header_row = ['No.']
    col_widths = [0.35*inch]
    
    for col in selected_columns:
        if col in column_config:
            header_row.append(column_config[col]['header'])
            col_widths.append(column_config[col]['width']*inch)
    
    # Table data
    data = [header_row]
    
    # Style for wrapping text in cells
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontSize=7,
        leading=8,
        wordWrap='CJK'
    )
    
    for idx, row in enumerate(results, 1):
        row_data = [str(idx)]
        
        for col in selected_columns:
            if col in column_config:
                value = row[column_config[col]['key']] or '-'
                
                # Wrap long text in Paragraph if needed
                if column_config[col]['wrap'] and len(str(value)) > 20:
                    row_data.append(Paragraph(str(value), cell_style))
                else:
                    row_data.append(str(value))
        
        data.append(row_data)
    
    # Create table
    table = Table(data, colWidths=col_widths, repeatRows=1)
    
    # Table style
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        
        # Body
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # No. column center
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('TOPPADDING', (0, 1), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Alternating rows
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Close cursor
    cursor.close()
    
    filename = f"monitoring_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )
    
@app.route('/tickets')
@login_required
def tickets():
    """Display all tickets"""
    tickets = load_tickets()
    
    # Filter options
    status_filter = request.args.get('status', '')
    section_filter = request.args.get('section', '')
    search_query = request.args.get('search', '').lower()
    
    # Apply filters
    filtered_tickets = tickets
    
    if status_filter:
        filtered_tickets = [t for t in filtered_tickets if t['status'] == status_filter]
    
    if section_filter:
        filtered_tickets = [t for t in filtered_tickets if t['section'] == section_filter]
    
    if search_query:
        filtered_tickets = [
            t for t in filtered_tickets 
            if search_query in t['ticket_no'].lower() 
            or search_query in t['title'].lower()
            or search_query in t['description'].lower()
        ]
    
    # Get unique sections for filter dropdown
    sections = list(set(t['section'] for t in tickets))
    sections.sort()
    
    return render_template(
        'tickets.html', 
        tickets=filtered_tickets,
        all_tickets=tickets,
        sections=sections,
        filters={
            'status': status_filter,
            'section': section_filter,
            'search': search_query
        }
    )

@app.route('/tickets/add', methods=['GET', 'POST'])
@login_required
def add_ticket():
    """Add new ticket"""
    if request.method == 'POST':
        tickets = load_tickets()
        
        new_ticket = {
            'ticket_no': generate_ticket_number(),
            'title': request.form['title'],
            'description': request.form['description'],
            'section': request.form['section'],
            'status': request.form['status'],
            'priority': request.form['priority'],
            'app_version': request.form.get('app_version', ''),
            'solved_version': request.form.get('solved_version', ''),
            'created_by': current_user.username,
            'created_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'notes': request.form.get('notes', ''),
            'history': [
                {
                    'action': 'Created',
                    'user': current_user.username,
                    'timestamp': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'details': 'Ticket created'
                }
            ]
        }
        
        tickets.append(new_ticket)
        save_tickets(tickets)
        
        log_action('create', 'tickets', new_ticket['ticket_no'], None, new_ticket['title'])
        
        flash(f'Ticket {new_ticket["ticket_no"]} created successfully!', 'success')
        return redirect(url_for('tickets'))
    
    return render_template('ticket_form.html', ticket=None)

@app.route('/tickets/edit/<ticket_no>', methods=['GET', 'POST'])
@login_required
def edit_ticket(ticket_no):
    """Edit existing ticket"""
    tickets = load_tickets()
    ticket = next((t for t in tickets if t['ticket_no'] == ticket_no), None)
    
    if not ticket:
        flash('Ticket not found!', 'danger')
        return redirect(url_for('tickets'))
    
    if request.method == 'POST':
        # Store old values for history
        old_status = ticket['status']
        old_section = ticket['section']
        
        # Update ticket
        ticket['title'] = request.form['title']
        ticket['description'] = request.form['description']
        ticket['section'] = request.form['section']
        ticket['status'] = request.form['status']
        ticket['priority'] = request.form['priority']
        ticket['app_version'] = request.form.get('app_version', '')
        ticket['solved_version'] = request.form.get('solved_version', '')
        ticket['notes'] = request.form.get('notes', '')
        ticket['updated_at'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Add to history
        changes = []
        if old_status != ticket['status']:
            changes.append(f"Status: {old_status} → {ticket['status']}")
        if old_section != ticket['section']:
            changes.append(f"Section: {old_section} → {ticket['section']}")
        
        if changes or request.form.get('history_note'):
            history_entry = {
                'action': 'Updated',
                'user': current_user.username,
                'timestamp': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'details': ', '.join(changes) if changes else request.form.get('history_note', 'Ticket updated')
            }
            ticket['history'].append(history_entry)
        
        save_tickets(tickets)
        
        log_action('update', 'tickets', ticket_no, old_status, ticket['status'])
        
        flash(f'Ticket {ticket_no} updated successfully!', 'success')
        return redirect(url_for('ticket_detail', ticket_no=ticket_no))
    
    return render_template('ticket_form.html', ticket=ticket)

@app.route('/tickets/<ticket_no>')
@login_required
def ticket_detail(ticket_no):
    """View ticket details"""
    tickets = load_tickets()
    ticket = next((t for t in tickets if t['ticket_no'] == ticket_no), None)
    
    if not ticket:
        flash('Ticket not found!', 'danger')
        return redirect(url_for('tickets'))
    
    return render_template('ticket_detail.html', ticket=ticket)

@app.route('/tickets/delete/<ticket_no>', methods=['POST'])
@login_required
@admin_required
def delete_ticket(ticket_no):
    """Delete ticket (admin only)"""
    tickets = load_tickets()
    ticket = next((t for t in tickets if t['ticket_no'] == ticket_no), None)
    
    if ticket:
        tickets = [t for t in tickets if t['ticket_no'] != ticket_no]
        save_tickets(tickets)
        
        log_action('delete', 'tickets', ticket_no, ticket['title'], None)
        
        flash(f'Ticket {ticket_no} deleted successfully!', 'success')
    else:
        flash('Ticket not found!', 'danger')
    
    return redirect(url_for('tickets'))

@app.route('/tickets/add_comment/<ticket_no>', methods=['POST'])
@login_required
def add_ticket_comment(ticket_no):
    """Add comment to ticket"""
    tickets = load_tickets()
    ticket = next((t for t in tickets if t['ticket_no'] == ticket_no), None)
    
    if not ticket:
        return jsonify({'success': False, 'message': 'Ticket not found'})
    
    comment = request.form.get('comment', '').strip()
    if not comment:
        return jsonify({'success': False, 'message': 'Comment cannot be empty'})
    
    history_entry = {
        'action': 'Comment',
        'user': current_user.username,
        'timestamp': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'details': comment
    }
    
    ticket['history'].append(history_entry)
    ticket['updated_at'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    save_tickets(tickets)
    
    return jsonify({'success': True, 'message': 'Comment added successfully'})

@app.route('/tickets/export/<ticket_no>')
@login_required
def export_ticket(ticket_no):
    """Export ticket as text format (similar to email format)"""
    tickets = load_tickets()
    ticket = next((t for t in tickets if t['ticket_no'] == ticket_no), None)
    
    if not ticket:
        flash('Ticket not found!', 'danger')
        return redirect(url_for('tickets'))
    
    # Create text format
    output = f"""Ticket No: {ticket['ticket_no']}
Sent to Wintec Group on {ticket['created_at'].split()[0]}

Title: {ticket['title']}
Section: {ticket['section']}
Status: {ticket['status']}
Priority: {ticket['priority']}
App Version: {ticket.get('app_version', 'N/A')}
Solved Version: {ticket.get('solved_version', 'N/A')}

Description:
{ticket['description']}

Notes:
{ticket.get('notes', 'N/A')}

History:
"""
    
    for entry in ticket['history']:
        output += f"\n[{entry['timestamp']}] {entry['user']} - {entry['action']}: {entry['details']}"
    
    # Return as downloadable file
    from io import BytesIO
    buffer = BytesIO()
    buffer.write(output.encode('utf-8'))
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='text/plain',
        as_attachment=True,
        download_name=f'{ticket_no}_export.txt'
    )

@app.route('/api/tickets/stats')
@login_required
def ticket_stats():
    """Get ticket statistics for dashboard"""
    tickets = load_tickets()
    
    stats = {
        'total': len(tickets),
        'solved': len([t for t in tickets if t['status'] == 'solved']),
        'pending': len([t for t in tickets if t['status'] == 'pending']),
        'testing': len([t for t in tickets if t['status'] == 'testing']),
        'by_section': {}
    }
    
    # Count by section
    for ticket in tickets:
        section = ticket['section']
        if section not in stats['by_section']:
            stats['by_section'][section] = {'total': 0, 'solved': 0, 'pending': 0}
        
        stats['by_section'][section]['total'] += 1
        if ticket['status'] == 'solved':
            stats['by_section'][section]['solved'] += 1
        elif ticket['status'] == 'pending':
            stats['by_section'][section]['pending'] += 1
    
    return jsonify(stats)    

# Error handlers
@app.errorhandler(404)
def not_found(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('500.html'), 500

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)